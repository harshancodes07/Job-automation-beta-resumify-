from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from backend.models.schemas import JobListing

HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LINK_FONT = Font(color="4A90D9", underline="single")

COLUMNS = [
    ("Title", 30),
    ("Company", 25),
    ("Location", 20),
    ("Posted Date", 14),
    ("Salary", 22),
    ("Source", 12),
    ("Apply URL", 40),
    ("Description", 60),
]


def export_to_excel(listings: list[JobListing]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Listings"

    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, job in enumerate(listings, start=2):
        ws.cell(row=row_idx, column=1, value=job.title)
        ws.cell(row=row_idx, column=2, value=job.company)
        ws.cell(row=row_idx, column=3, value=job.location)
        ws.cell(row=row_idx, column=4, value=job.posted_date)
        ws.cell(row=row_idx, column=5, value=job.salary_range)
        ws.cell(row=row_idx, column=6, value=job.source)

        url_cell = ws.cell(row=row_idx, column=7, value=job.apply_url)
        if job.apply_url:
            url_cell.hyperlink = job.apply_url
            url_cell.font = LINK_FONT

        desc_cell = ws.cell(row=row_idx, column=8, value=job.description[:500])
        desc_cell.alignment = Alignment(wrap_text=True)

    ws.auto_filter.ref = ws.dimensions

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
